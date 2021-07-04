# coding: UTF-8
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
import time
import json

# ============= main def =====================================

def generate_common_race_list(data: str):
  newList = []
  isPassedCourse = False

  for index,elm in enumerate(data):

    if isPassedCourse == False:
      # 最初を通過してからコースの長さ(H1800m)が出てくるまでスキップ
      if index == 0:
        newList.append(elm)
      else:
        if 'H' in elm:
          newList.append(elm)
          isPassedCourse = True
    else:
      if not "風" in elm and not "波" in elm:
        if "m" in elm:
          if "cm" in elm:
            newList.append(elm.replace("cm",""))
          else:
            newList.append(elm.replace("m",""))
        else:
          newList.append(elm)

  
  return newList

def generate_race_result_list(datalist: list,indexElm: int):
  allList = []
  for counter in range(6):
    resultlist = datalist[indexElm+counter].split()
    generatedResultList = []
    playerName = []
    dotCount = 0

    # 各選手のレース結果と情報(横列)
    for dx,result in enumerate(resultlist):
      if is_int(result):
        generatedResultList.append(result)
      elif is_float(result):
        generatedResultList.append(result)
      elif dx == len(resultlist)-1:
        if result == '.':
          generatedResultList.append(".  .")
        else:
          generatedResultList.append(result)
      else:
        if result == '.':
          continue
        elif 'K' in result or 'S' in result or 'F' in result or 'L1' in result:
          generatedResultList.append(result)
        else:
          playerName.append(result)

      if dx == 9:
        generatedResultList.insert(0, ''.join(playerName))

    allList.append(generatedResultList)

  return allList

def is_int(s):  
    p = '[-+]?\d+'
    return True if re.fullmatch(p, s) else False

def is_float(s):  
    p = '[-+]?(\d+\.?\d*|\.\d+)([eE][-+]?\d+)?'
    return True if re.fullmatch(p, s) else False

def create_Excel_Data_from_Local_data(columnCount: int, index: int, data:str, sheet, txtFile, datalist: list):
  indexElm = 0
  roundDict = {"着順":" ","艇番":" ","登録番号":" ","選手名":" ","オッズ":" ","体重":" ","ﾓｰﾀｰ":" ","ﾎﾞｰﾄ":" ","展示":" ","進入":" ","ST":" ","ﾚｰｽﾀｲﾑ":" ","ラウンド":" ","コース(m)":" ","天気":" ","風向":" ","風速(m)":" ","波(cm)":" ","日付":" ","開催場所":" ","大会名":" ","SG":" ","G1":" ","G2":" ","G3":" ","一般":" ",}
  
  #　①[選手名]までローカル配列から取得、[オッズ][体重]は最初は空のstrを入れておく、（SGより右も同様）
  # これらを配列に挿入してcsvに入れる準備をする。6回回ったら終了
  if 'R' in data and index > 2:
    generatedData = generate_common_race_list(data.split())

    roundDict.update(
      {
        "ラウンド" : int(generatedData[0][:-1]),
        "コース(m)" : generatedData[1][:-1],
        "天気" : generatedData[2],
        "風向" : generatedData[3],
        "風速(m)" : int(generatedData[4]),
        "波(cm)" : int(generatedData[5]),
        "日付" : txtFile[3:],
      }
    )
    indexElm = index + 3
  
  # データ項目名を取得
  if index == 0:
    # セルへ書き込む
    indexListCount = 1
    for  headElm in roundDict.keys():
      sheet.cell(row=1,column=indexListCount).value = headElm
      indexListCount += 1

  # レースの結果分を取得(縦列)
  if indexElm != 0:
    columnCountMin = columnCount
    for raceResult in generate_race_result_list(datalist=datalist, indexElm=indexElm):
      intrusion = ".  ."
      st = ".  ."
      raceTime = ".  ."

      # indexが7以降要素があるかどうか分岐する
      if raceResult[7:8]:
        intrusion = raceResult[7]
      if raceResult[8:9]:
        st = raceResult[8]
      if raceResult[9:10]:
        raceTime = raceResult[9]

      roundDict.update(
        {
          "着順" : raceResult[1],
          "艇番" : int(raceResult[2]),
          "登録番号" : int(raceResult[3]),
          "選手名" : raceResult[0],
          "ﾓｰﾀｰ" : int(raceResult[4]),
          "ﾎﾞｰﾄ" : int(raceResult[5]),
          "展示" : raceResult[6],
          "進入" : intrusion,
          "ST" : st,
          "ﾚｰｽﾀｲﾑ" : raceTime,
        }
      )

      # 横列のindex
      cellIndex = 1
      # columnCountを固定させて、roundDictから艇番を取得して、艇番順に並び替える
      boatNumber = columnCountMin + int(roundDict["艇番"]) - 1

      for cellValue in roundDict.values():
        sheet.cell(row=boatNumber,column=cellIndex).value = cellValue
        cellIndex += 1
      columnCount += 1
  return columnCount

def generate_one_day_race_result_list(jsonData):
  txtFile = f'./local_boat_data/{jsonData["txtFile"]}.TXT'
  f = open(txtFile, 'r', encoding='shift_jis')
  datalist = f.readlines()
  f.close()

  placeTour = []
  columnCount = 2

  for index,data in enumerate(datalist):
    if '［成績］' in data:
      c = data.find("［")
      placeTour.append(re.sub(r"[\u3000 \t \n]","",data[:c]))
      placeTour.append(re.sub(r"[\u3000 \t \n]","",datalist[index+4]))

    columnCount = create_Excel_Data_from_Local_data(
      columnCount=columnCount,
      index=index,
      data=data,
      sheet=sheet,
      txtFile=jsonData["txtFile"],
      datalist=datalist
    )
  
  return placeTour

# =========== scraping def =============================================

def requested_url(reqUrl: str):
  res = requests.get(reqUrl)
  res.encoding = res.apparent_encoding

  return BeautifulSoup(res.text, "lxml")

def raceListByRound(roundTag: str):
  detailRound = requested_url(roundTag)
  roundPage = detailRound.find("ul", class_="tab3_tabs")
  spans = roundPage.find_all("a")

  oddAndInfoUrlList: str = []

  for span in spans:
    spanTag = "https://www.boatrace.jp" + span.get("href")
    oddAndInfoUrlList.append(spanTag)

  oddUrl = oddAndInfoUrlList[0].replace("odds3t","oddstf")

  oddPage = requested_url(oddUrl)
  oddPoints = oddPage.find_all("td" ,class_="oddsPoint")

  oddPointsList = []

  for index,point in enumerate(oddPoints):
    if index < 6:
      oddPointsList.append(point.text)

  detailInfoPage = requested_url(oddAndInfoUrlList[1])
  weightPage = detailInfoPage.find("table",class_="is-w748")
  weightTable = weightPage.find_all("td")

  weightList = []

  for weight in weightTable:
    if "kg" in weight.text:
      weightList.append(weight.text)

  return oddPointsList,weightList

def scrapingOnetournament(url: str, requestCount: int):
  html = requested_url(url)

  table1 = html.find("div", class_="table1")

  roundTagList = []
  resultList = []
  # oddとinfoのurl取得
  for a in table1.find_all('td', class_="is-fs14 is-fBold"):
    roundTag = "https://www.boatrace.jp" + a.find("a").get("href")
    roundTagList.append(roundTag)

  for raceRound in roundTagList:
    resultList.append(raceListByRound(raceRound))
    time.sleep(2)

  columnCount = 2 + 72 * requestCount

  # 1Race、１ループ
  for index,result in enumerate(resultList):
    odds = result[0]
    weights = result[1]

    for odd in odds:
      # その位置の艇番
      sheet.cell(row=columnCount,column=5).value = odd
      columnCount += 1

    columnCount -= 6

    for weight in weights:
      # その位置の艇番
      sheet.cell(row=columnCount,column=6).value = weight[:-2]
      columnCount += 1

# =========== other def =============================================

def generate_other_boat_data(placeTour: list,reservedRaceRank: list):
  raceRankDict = {"SG" : [1,0,0,0,0], "G1" : [0,1,0,0,0], "G2" : [0,0,1,0,0], "G3" : [0,0,0,1,0], "GE" : [0,0,0,0,1]}
  existPlace = False
  tourCount = 0
  tourSum = 2

  for va in placeTour:
    for i in range(tourSum, tourSum+72):
      if tourCount >= 2:
        i = i - (2 * tourCount) + 2
      if existPlace == False:
        sheet.cell(row=i,column=20).value = va
      else:
        sheet.cell(row=i,column=21).value = va
        currentRaceRank = reservedRaceRank[tourCount]
        for index in range(len(raceRankDict[currentRaceRank])):
          sheet.cell(row=i,column=22+index).value = raceRankDict[currentRaceRank][index]

    if existPlace == False:
      existPlace = True
    else:
      tourCount += 1
      existPlace = False
      tourSum = (tourCount * 74)


# =========== main function =============================================

jsonFile = "./input_boat_info.json"
fd = open(jsonFile, mode='r')
jsonData = json.load(fd)
fd.close()

for i in range(len(jsonData)):
  stringIndex = str(i+1)
  urlList = jsonData[stringIndex]["urlList"]

  if len(urlList) != len(jsonData[stringIndex]["raceRankList"]):
    print("urlListとraceRankListの数が一致していません")
    break

  exfile = jsonData[stringIndex]["excelFile"]
  excelFile = f'./excel_data/{exfile}.xlsx'
  book = openpyxl.load_workbook(excelFile)
  sheet = book.worksheets[0]

  print(f"{exfile}分のボートデータをexcelに書き込み中")

  placeTour = generate_one_day_race_result_list(jsonData=jsonData[stringIndex])

  ccp = 0
  for link in urlList:
    scrapingOnetournament(url=link,requestCount=ccp)
    ccp += 1
    print(f'スクレイピング進捗 : {ccp}/{len(urlList)}')

  generate_other_boat_data(placeTour=placeTour, reservedRaceRank=jsonData[stringIndex]["raceRankList"])

  # 保存する
  book.save(excelFile)
  print(f"{exfile}分の書き込みが完了しました")
  print("=============================================")