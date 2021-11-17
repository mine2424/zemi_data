# coding: UTF-8
import datetime
import openpyxl
import json
from openpyxl.workbook import workbook
import generate_boat_race
import scraping_boat_data

# =========== other def =============================================


def generate_other_boat_data(placeTour: list, reservedRaceRank: list, roundIndexList: list):
    raceRankDict = {
        "SG": [1, 0, 0, 0, 0],
        "G1": [0, 1, 0, 0, 0],
        "G2": [0, 0, 1, 0, 0],
        "G3": [0, 0, 0, 1, 0],
        "GE": [0, 0, 0, 0, 1],
    }
    existPlace = False
    tourCount = 0
    tourSum = 2

    # 個々のfor,if文等のセクションでどのような処理が行われているか見ること。
    # 各大会の「地名」「大会名」がワンセットのリストを分解して（回して）いる。
    for index, va in enumerate(placeTour):
        # 最後のカラムのインデックスからその先の次のレースの分まで回転する。
        roundIndex = int(index / 2)
        addedTourSum = tourSum + (int(roundIndexList[roundIndex])*6)
        for i in range(tourSum, addedTourSum):
            if tourCount >= 2:
                i = i - (2 * tourCount) + 2
            if existPlace == False:
                sheet.cell(row=i, column=21).value = va
            else:
                sheet.cell(row=i, column=22).value = va
                currentRaceRank = reservedRaceRank[tourCount]
                for j in range(len(raceRankDict[currentRaceRank])):
                    sheet.cell(
                        row=i,
                        column=23 + j
                    ).value = raceRankDict[currentRaceRank][j]

        if existPlace == False:
            existPlace = True
        else:
            tourCount += 1
            existPlace = False

            if roundIndex == 0:
                tourSum = addedTourSum
            else:
                tourSum = addedTourSum + 2


# =========== main function =============================================

# jsonFile = "../input_boat_info.json"
# fd = open(jsonFile, mode='r')
# jsonData = json.load(fd)
# fd.close()
#
# for i in range(len(jsonData)):
#     stringIndex = str(i+1)
#     urlList = jsonData[stringIndex]["urlList"]
#     roundIndexList = jsonData[stringIndex]["roundIndexList"]
#
#     if len(urlList) != len(jsonData[stringIndex]["raceRankList"]):
#         print("urlListとraceRankListの数が一致していません")
#         break
#
#     initializedSheet = openpyxl.Workbook()
#     exfile = jsonData[stringIndex]["excelFile"]
#     excelFile = f'../excel_data/{exfile}.xlsx'
#     initializedSheet.save(excelFile)
#     book = openpyxl.load_workbook(excelFile)
#     sheet = book.worksheets[0]
#
#     print(f"{exfile}分のボートデータをexcelに書き込み中")
#     placeTour = generate_boat_race.generate_one_day_race_result_list(
#         sheet=sheet,
#         jsonData=jsonData[stringIndex],
#     )
#
#     # ccp = 0
#     # columnCount = 2
#     # for index,link in enumerate(urlList):
#     #   columnCount = scraping_boat_data.scrapingOnetournament(sheet=sheet,url=link,columnCount=columnCount,roundCount=int(roundIndexList[index]))
#     #   ccp += 1
#     #   print(f'スクレイピング進捗 : {ccp}/{len(urlList)}')
#
#     generate_other_boat_data(
#         placeTour=placeTour,
#         reservedRaceRank=jsonData[stringIndex]["raceRankList"],
#         roundIndexList=roundIndexList,
#     )
#
#     # 保存する
#     book.save(excelFile)
#     print(f"{exfile}分の書き込みが完了しました")
#     print("=============================================")
#     print(" ")


print(scraping_boat_data.generate_all_tournament_url(
    datetime.date(2018, 1, 1),
    datetime.date(2020, 12, 31),
))
