# coding: UTF-8
import openpyxl

# =========== other def =============================================

# round_list_by_day: [12, 12, 12, 12, 12, 13, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
def generate_other_boat_data(
    pre_tour_count: int, 
    tour_place_list: list, 
    tour_name_list: list, 
    reservedRaceRank: list, 
    roundIndexList: list, 
    time_zone_list: list, 
    tour_series_list: list, 
    sheet
):
    raceRankDict = {
        "SG": [1, 0, 0, 0, 0],
        "G1": [0, 1, 0, 0, 0],
        "G2": [0, 0, 1, 0, 0],
        "G3": [0, 0, 0, 1, 0],
        "GE": [0, 0, 0, 0, 1],
    }

    time_zone_dict = {
        "none": 0,
        "nighter": 1,
        "summer": 2,
        "morning": 3,
    }

    tour_series_dict = {
        "none": 0,
        "venus": 1,
        "lady": 2,
        "rookie": 3,
    }

    tourCount = 0
    tourSum = pre_tour_count
    
    if len(roundIndexList) == 0:
        for i in range(len(roundIndexList)):
            if roundIndexList[i] == -1:
                roundIndexList.pop(i)

    # 個々のfor,if文等のセクションでどのような処理が行われているか見ること。
    # 各大会の「地名」「大会名」がワンセットのリストを分解して（回して）いる。
    print(f"tour_place_list: {tour_place_list},roundIndexList: {roundIndexList}, len(tour_place_list): {len(tour_place_list)}, len(roundIndexList): {len(roundIndexList)}")
    for index in range(len(tour_place_list)):
        # 最後のカラムのインデックスからその先の次のレースの分まで回転する。
        if roundIndexList[index] != -1:
            addedTourSum = tourSum + (int(roundIndexList[index]) * 6)
            print(f"roundIndexList[index]: {roundIndexList[index]}, add - pre: {addedTourSum-tourSum}")
            for i in range(tourSum, addedTourSum):
                if tourCount >= 2:
                    i = i - (2 * tourCount) + 2
                # 開催場所
                sheet.cell(row=i, column=22).value = tour_place_list[index]
                # 大会名
                sheet.cell(row=i, column=23).value = tour_name_list[index]
                
                currentRaceRank = reservedRaceRank[tourCount]
                if currentRaceRank != None:
                    for j in range(len(raceRankDict[currentRaceRank])):
                        sheet.cell(
                            row=i,
                            column=24 + j
                        ).value = raceRankDict[currentRaceRank][j]
                # 時間帯、シリーズのダミー変数をExcelに挿入
                current_time_zone = time_zone_list[tourCount]
                current_tour_series = tour_series_list[tourCount]
                sheet.cell(
                    row=i, column=29).value = time_zone_dict[current_time_zone]
                sheet.cell(
                    row=i, column=30).value = tour_series_dict[current_tour_series]

            tourCount += 1
            if index == 0:
                tourSum = addedTourSum
            else:
                tourSum = addedTourSum + 2


# TODO:保存後のexcelの空白を削除する機能を実装する
def delete_empty_rows_in_excel(file_name: str):
    wb = openpyxl.load_workbook(f'../excel_data/{file_name}.xlsx')
    ws = wb.worksheets[0]
    sheet_max = ws.max_row + 1
    # 最終行から逆ループ
    for i in reversed(range(1, sheet_max)):
        # A列が None だったら
        if ws.cell(row=i, column=1).value == None and type(ws.cell(row=i, column=21).value) is str:
            ws.delete_rows(i)
            print(f'delete index: {i}')
    
    wb.save(f'../excel_data/{file_name}.xlsx')
    return wb, ws


# TODO:保存後のexcelの空白を削除する機能を実装する
def delete_empty_no_len_rows_in_excel(file_name: str):
    wb = openpyxl.load_workbook(f'../excel_data/{file_name}.xlsx')
    ws = wb.worksheets[0]
    sheet_max = ws.max_row + 1
    # 最終行から逆ループ
    for i in reversed(range(1, sheet_max)):
        GetValue1 = ws.cell(row=i, column=1).value

        #A列が None だったら
        if GetValue1 == None:
            #行削除
            ws.delete_rows(i)
        else:
            #A列の空白文字を削除
            GetValue2 = GetValue1.strip()
    
            #セルの文字数が 0 だったら
            if len(GetValue2) == 0:
                #行削除
                ws.delete_rows(i)
    
    wb.save(f'../excel_data/{file_name}.xlsx')
    return wb, ws
